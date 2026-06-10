from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

AZUL = "1E3A5F"; AZUL2 = "2E75B6"; FILL_SUB = "D5E8F0"; FILL_ALT = "F2F6FA"
VERDE = "1E7B34"; ROJO = "B02A2A"; AMAR = "FFF2B2"
WHITE = "FFFFFF"

FONT = "Arial"
f_title = Font(name=FONT, size=18, bold=True, color=AZUL)
f_sub   = Font(name=FONT, size=11, italic=True, color="555555")
f_h     = Font(name=FONT, size=11, bold=True, color=WHITE)
f_b     = Font(name=FONT, size=10, bold=True)
f_n     = Font(name=FONT, size=10)
f_blue  = Font(name=FONT, size=10, color="0000FF", bold=True)   # inputs
f_green = Font(name=FONT, size=10, color="008000", bold=True)   # cross-sheet links
f_red   = Font(name=FONT, size=10, bold=True, color=ROJO)
f_grn   = Font(name=FONT, size=10, bold=True, color=VERDE)
f_big   = Font(name=FONT, size=14, bold=True, color=AZUL)

fill_head = PatternFill("solid", fgColor=AZUL)
fill_sub  = PatternFill("solid", fgColor=FILL_SUB)
fill_alt  = PatternFill("solid", fgColor=FILL_ALT)
fill_amar = PatternFill("solid", fgColor=AMAR)
fill_band = PatternFill("solid", fgColor=AZUL2)

thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
right  = Alignment(horizontal="right", vertical="center")

CLP = '#,##0 "CLP";(#,##0) "CLP";"-"'
CLP0 = '#,##0;(#,##0);"-"'
PCT = '0%'
NUM = '#,##0'

wb = Workbook()

# =====================================================================
#  HOJA 3 (build first to reference): CALCULADORA DE AHORRO
# =====================================================================
cal = wb.active
cal.title = "Calculadora de Ahorro"
cal.sheet_view.showGridLines = False
for col, w in {"A":3,"B":40,"C":16,"D":16,"E":16,"F":16,"G":3}.items():
    cal.column_dimensions[col].width = w

def put(ws, cell, val, font=f_n, fill=None, al=None, fmt=None, bd=True):
    c = ws[cell]; c.value = val; c.font = font
    if fill: c.fill = fill
    if al: c.alignment = al
    if fmt: c.number_format = fmt
    if bd: c.border = border
    return c

cal["B2"] = "CALCULADORA DE AHORRO — CONTROL DE HORAS EXTRA"; cal["B2"].font = f_title
cal["B3"] = "Syncro Red · Celdas azules = editables por Finanzas · Negras = fórmulas automáticas"; cal["B3"].font = f_sub

# --- Sección A: dotación ---
cal["B5"] = "A. PARÁMETROS DE DOTACIÓN"; cal["B5"].font = f_h; cal["B5"].fill = fill_head
for c in ("C5","D5","E5","F5"): cal[c].fill = fill_head
r = 6
rows_A = [
    ("Sueldo mensual — Maquinista", 1989000, f_blue, CLP),
    ("Sueldo mensual — Ayudante", 980000, f_blue, CLP),
    ("Jornada legal (horas/semana)", 42, f_blue, NUM),
    ("Recargo legal hora extra", 1.5, f_blue, '0.0"x"'),
]
for label, val, fnt, fmt in rows_A:
    put(cal, f"B{r}", label, f_n, al=left)
    put(cal, f"C{r}", val, fnt, al=right, fmt=fmt)
    r += 1
cal["C6"].comment = Comment("Sueldo base mensual maquinista (dato real).", "Syncro Red")
cal["C8"].comment = Comment("Ley 40 horas: 42 h/sem vigente 2026.", "Syncro Red")
# valores HE (fórmulas)
put(cal, "B10", "Valor hora extra — Maquinista", f_n, al=left)
put(cal, "C10", "=C6*(7/(30*C8))*C9", f_n, al=right, fmt=CLP)
put(cal, "B11", "Valor hora extra — Ayudante", f_n, al=left)
put(cal, "C11", "=C7*(7/(30*C8))*C9", f_n, al=right, fmt=CLP)
put(cal, "B12", "Valor HE de DOTACIÓN (promedio maq+ayud)", f_b, fill=fill_amar, al=left)
put(cal, "C12", "=AVERAGE(C10:C11)", f_b, fill=fill_amar, al=right, fmt=CLP)
cal["C12"].comment = Comment("Promedio porque cada servicio lleva 1 maquinista + 1 ayudante. Fórmula legal chilena con recargo 50%.", "Syncro Red")

# --- Sección B: escenarios ---
cal["B14"] = "B. ESCENARIOS DE FUGA Y AHORRO"; cal["B14"].font = f_h; cal["B14"].fill = fill_head
for c in ("C14","D14","E14","F14"): cal[c].fill = fill_head
hdr = ["Concepto", "Conservador", "Esperado", "Optimista"]
for i, h in enumerate(hdr):
    cc = f"{get_column_letter(2+i)}15"
    put(cal, cc, h, f_h, fill=fill_head, al=center)

def row3(ws, rr, label, vals, fnt, fmt, fill=None, lblfont=None):
    put(ws, f"B{rr}", label, lblfont or f_n, al=left, fill=fill)
    for i, v in enumerate(vals):
        put(ws, f"{get_column_letter(3+i)}{rr}", v, fnt, al=right, fmt=fmt, fill=fill)

# inputs
row3(cal, 16, "Horas extra base / mes (red completa)", [1300, 1550, 1800], f_blue, NUM)
row3(cal, 17, "% incremento evitable (mala asignación)", [0.30, 0.35, 0.40], f_blue, PCT)
row3(cal, 18, "% de captura del incremento (efic. del sistema)", [0.50, 0.60, 0.70], f_blue, PCT)
cal["C16"].comment = Comment("Rango entregado por la operación: 1.300–1.800 HE/mes a nivel de red.", "Syncro Red")
cal["C18"].comment = Comment("Fracción del incremento evitable que el módulo logra eliminar. Conservador.", "Syncro Red")
# calcs
row3(cal, 19, "Gasto base HE / mes", ["=C16*$C$12","=D16*$C$12","=E16*$C$12"], f_n, CLP)
row3(cal, 20, "Gasto base HE / año", ["=C19*12","=D19*12","=E19*12"], f_n, CLP)
row3(cal, 21, "Incremento evitable (HE / mes)", ["=C16*C17","=D16*D17","=E16*E17"], f_n, NUM)
row3(cal, 22, "Fuga evitable / mes", ["=C21*$C$12","=D21*$C$12","=E21*$C$12"], f_red, CLP, fill=fill_alt, lblfont=f_b)
row3(cal, 23, "Fuga evitable / año", ["=C22*12","=D22*12","=E22*12"], f_red, CLP, fill=fill_alt, lblfont=f_b)
row3(cal, 24, "AHORRO ANUAL (fuga evitable × captura)", ["=C23*C18","=D23*D18","=E23*E18"], f_grn, CLP, fill=fill_sub, lblfont=f_b)
row3(cal, 25, "Ahorro mensual", ["=C24/12","=D24/12","=E24/12"], f_grn, CLP)

# --- Sección C: referencias e inversión ---
cal["B27"] = "C. REFERENCIAS DE VALOR E INVERSIÓN"; cal["B27"].font = f_h; cal["B27"].fill = fill_head
for c in ("C27","D27","E27","F27"): cal[c].fill = fill_head
put(cal, "B28", "Validación de mercado (caso Merval)", f_n, al=left)
put(cal, "C28", 7000000, f_blue, al=right, fmt=CLP)
cal["C28"].comment = Comment("Precedente: sistema inferior de otra red transado en CLP 7M. Referencia de piso, no un cobro.", "Syncro Red")
put(cal, "B29", "Desarrollo externo equivalente — rango", f_n, al=left)
put(cal, "C29", 25000000, f_blue, al=right, fmt=CLP)
put(cal, "D29", 60000000, f_blue, al=right, fmt=CLP)
put(cal, "B30", "Payback vs fuga evitable (meses) — ref. CLP 7M", f_b, al=left)
row3(cal, 30, "Payback vs fuga evitable (meses) — ref. CLP 7M", ["=$C$28/C22","=$C$28/D22","=$C$28/E22"], f_b, '0.0', lblfont=f_b)
cal["B31"] = "Lectura: el costo de referencia (CLP 7M) se recupera en menos de ~1,5 meses de fuga evitable."
cal["B31"].font = f_sub
cal.freeze_panes = "B6"

# =====================================================================
#  HOJA 1: RESUMEN EJECUTIVO
# =====================================================================
res = wb.create_sheet("Resumen Ejecutivo", 0)
res.sheet_view.showGridLines = False
for col, w in {"A":3,"B":34,"C":22,"D":22,"E":22,"F":3}.items():
    res.column_dimensions[col].width = w

res["B2"] = "SYNCRO RED — RESUMEN EJECUTIVO PARA LA GERENCIA"; res["B2"].font = f_title
res["B3"] = "Control de presupuesto de horas extra · Lectura rápida para Gerencia, Jefe de Operaciones, IL y SL"; res["B3"].font = f_sub

res["B5"] = "EL MENSAJE EN UNA FRASE"; res["B5"].font = f_h; res["B5"].fill = fill_head
for c in ("C5","D5","E5"): res[c].fill = fill_head
res.merge_cells("B6:E8")
res["B6"] = ("No es una app de horarios: es una herramienta de control de presupuesto que detiene una fuga "
             "proyectada en horas extra, construida por personal de la propia vía, y que se paga sola en semanas.")
res["B6"].font = Font(name=FONT, size=11, bold=True, color=AZUL)
res["B6"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
for rr in range(6,9):
    for cc in "BCDE":
        res[f"{cc}{rr}"].fill = fill_sub
        res[f"{cc}{rr}"].border = border

# KPIs
res["B10"] = "CIFRAS CLAVE (escenario esperado)"; res["B10"].font = f_h; res["B10"].fill = fill_head
for c in ("C10","D10","E10"): res[c].fill = fill_head
kpis = [
    ("Valor hora extra de dotación", "='Calculadora de Ahorro'!C12", CLP, f_green),
    ("Fuga evitable / año", "='Calculadora de Ahorro'!D23", CLP, f_red),
    ("Ahorro anual estimado", "='Calculadora de Ahorro'!D24", CLP, f_grn),
    ("Payback (referencia CLP 7M)", "='Calculadora de Ahorro'!D30", '0.0 "meses"', f_b),
]
r = 11
for label, formula, fmt, fnt in kpis:
    put(res, f"B{r}", label, f_n, al=left)
    res.merge_cells(f"C{r}:E{r}")
    put(res, f"C{r}", formula, fnt, al=center, fmt=fmt)
    for cc in "DE": res[f"{cc}{r}"].border = border
    r += 1

# recomendación
res["B16"] = "QUÉ PEDIMOS (modelo ganar-ganar)"; res["B16"].font = f_h; res["B16"].fill = fill_head
for c in ("C16","D16","E16"): res[c].fill = fill_head
asks = [
    "Quedar a cargo del sistema (custodia técnica) — la PI es nuestra; la empresa opera bajo licencia de uso interno.",
    "Compensación monetaria por responsabilidad + resultados + criticidad, autofinanciada con el ahorro.",
    "Régimen 50/50 durante el desarrollo (50% turno / 50% desarrollo): respaldo operativo total, riesgo cero.",
    "Reconocimiento de autoría y, a futuro y por mérito, valorar una transición a administración/desarrollo.",
]
r = 17
for a in asks:
    res.merge_cells(f"B{r}:E{r}")
    cell = res[f"B{r}"]; cell.value = "•  " + a; cell.font = f_n
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for cc in "BCDE":
        if r % 2:
            res[f"{cc}{r}"].fill = fill_alt
        res[f"{cc}{r}"].border = border
    res.row_dimensions[r].height = 30
    r += 1
res.row_dimensions[6].height = 22

# =====================================================================
#  HOJA 2: COMPARATIVA DE OPCIONES
# =====================================================================
comp = wb.create_sheet("Comparativa de Opciones", 1)
comp.sheet_view.showGridLines = False
for col, w in {"A":3,"B":30,"C":30,"D":28,"E":26,"F":3}.items():
    comp.column_dimensions[col].width = w

comp["B2"] = "MATRIZ COMPARATIVA — 3 OPCIONES SOBRE LA MESA"; comp["B2"].font = f_title
comp["B3"] = "Valorización fácil de leer para Gerencia, Jefe de Operaciones, IL y SL"; comp["B3"].font = f_sub

headers = ["Criterio", "Syncro Red (nosotros)", "“Tren OS” (réplica TI)", "Desarrollo externo"]
for i, h in enumerate(headers):
    cc = f"{get_column_letter(2+i)}5"
    put(comp, cc, h, f_h, fill=fill_head, al=center)
comp.row_dimensions[5].height = 30

data = [
    ("Estado", "Ya construido y funcional", "A replicar / adaptar", "A iniciar de cero"),
    ("Conocimiento operativo", "Nativo (personal de vía)", "Importado de otra red", "Nulo (levantamiento largo)"),
    ("Control de presupuesto de HE", "Motor activo", "No (solo muestra turnos)", "Solo si se paga aparte"),
    ("Normativa ferroviaria embebida", "Codificada y validada en terreno", "Lógica de otra red", "Requiere levantamiento"),
    ("Costo directo", "Custodia + compensación (autofinanciada)", "“Gratis” aparente", "CLP 25M – 60M"),
    ("Tiempo a producción", "Semanas (hardening)", "Meses inciertos", "6 – 12 meses"),
    ("Adopción del operador", "Alta (hecho por pares)", "Baja", "Incierta"),
    ("Quién sufre el error de asignación", "Nosotros, en la vía", "TI, fuera de la línea", "Proveedor ajeno"),
    ("Mejora continua", "En horas, desde la operación", "Cola de TI", "Por hora, costo externo"),
    ("Impacto financiero anual", "Ahorro CLP 28M–73M", "Mantiene/aumenta la fuga", "Costo + ahorro incierto"),
]
r = 6
for crit, a, b, c in data:
    fillrow = fill_alt if (r % 2 == 0) else None
    put(comp, f"B{r}", crit, f_b, al=left, fill=fillrow)
    put(comp, f"C{r}", a, f_grn, al=left, fill=fillrow)
    put(comp, f"D{r}", b, f_red, al=left, fill=fillrow)
    put(comp, f"E{r}", c, f_n, al=left, fill=fillrow)
    comp.row_dimensions[r].height = 26
    r += 1

# valorización final
put(comp, f"B{r}", "VALORIZACIÓN", f_h, fill=fill_head, al=left)
put(comp, f"C{r}", "Activo entregado + ahorro recurrente", f_h, fill=fill_head, al=left)
put(comp, f"D{r}", "Fuga intacta (costo oculto)", f_h, fill=fill_head, al=left)
put(comp, f"E{r}", "Mayor desembolso, más lento", f_h, fill=fill_head, al=left)
comp.row_dimensions[r].height = 26
r += 2
comp.merge_cells(f"B{r}:E{r}")
comp[f"B{r}"] = ("Conclusión: el “gratis” de TI es el escenario más caro — un clon que muestra turnos pero asigna mal "
                 "mantiene intacta la fuga de CLP 56M–104M al año.")
comp[f"B{r}"].font = Font(name=FONT, size=10, bold=True, color=AZUL)
comp[f"B{r}"].alignment = left
for cc in "BCDE":
    comp[f"{cc}{r}"].fill = fill_sub; comp[f"{cc}{r}"].border = border
comp.row_dimensions[r].height = 36

from openpyxl.workbook.properties import CalcProperties
wb.calculation = CalcProperties(fullCalcOnLoad=True)

wb.save("Comparativa_y_Ahorro_Syncro_Red.xlsx")
print("OK -> Comparativa_y_Ahorro_Syncro_Red.xlsx")

# --- verificacion manual de las formulas clave ---
maq, ay, jor, rec = 1989000, 980000, 42, 1.5
he_maq = maq*(7/(30*jor))*rec
he_ay  = ay*(7/(30*jor))*rec
he_dot = (he_maq+he_ay)/2
print(f"HE maq={he_maq:,.0f}  HE ay={he_ay:,.0f}  HE dotacion={he_dot:,.0f}")
for name, base, incr, capt in [("Conserv",1300,0.30,0.50),("Esperado",1550,0.35,0.60),("Optim",1800,0.40,0.70)]:
    fuga_mes = base*incr*he_dot
    fuga_ano = fuga_mes*12
    ahorro = fuga_ano*capt
    payback = 7000000/fuga_mes
    print(f"{name:9s} fuga/ano={fuga_ano/1e6:5.1f}M  ahorro={ahorro/1e6:5.1f}M  payback={payback:.1f}m")
