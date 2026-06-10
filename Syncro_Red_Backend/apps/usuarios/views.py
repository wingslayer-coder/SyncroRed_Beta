import secrets
import datetime
import logging
from math import floor
from rest_framework import viewsets, views, status
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.throttling import ScopedRateThrottle
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
from django.db import transaction
from .models import Usuario, RegistroOperativo, AusenciaTemporal
from .serializers import UsuarioSerializer, RegistroOperativoSerializer, AusenciaTemporalSerializer
from .permissions import IsAdminOrStaff, IsJefaturaOSuperior, IsPropioUsuarioOJefatura

logger = logging.getLogger(__name__)


class LoginRateThrottle(ScopedRateThrottle):
    scope = 'login'


class LoginRutView(views.APIView):
    permission_classes = []
    authentication_classes = []
    throttle_classes = [LoginRateThrottle]

    def post(self, request):
        rut = request.data.get('rut')
        password = request.data.get('password')

        if not rut or not password:
            return Response({'error': 'RUT y contraseña son requeridos'}, status=400)

        try:
            user = Usuario.objects.get(rut=rut)
        except Usuario.DoesNotExist:
            logger.warning('Login fallido: RUT no encontrado')
            return Response({'error': 'Credenciales inválidas'}, status=401)

        if not user.check_password(password):
            logger.warning('Login fallido: contraseña incorrecta para RUT existente')
            return Response({'error': 'Credenciales inválidas'}, status=401)

        refresh = RefreshToken.for_user(user)
        return Response({
            'access': str(refresh.access_token),
            'refresh': str(refresh),
            'usuario': {
                'rut': user.rut,
                'nombre': user.nombre,
                'apellido': user.apellido,
                'cargo': user.cargo,
                'is_active': user.is_active,
                'is_staff': user.is_staff,
                'must_change_password': user.must_change_password,
            }
        })


class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer
    permission_classes = [IsJefaturaOSuperior]

    @action(detail=True, methods=['post'], permission_classes=[IsAdminOrStaff])
    def reset_password(self, request, pk=None):
        """Admin resetea la contraseña de un usuario y genera una temporal."""
        user = self.get_object()
        temp_password = secrets.token_urlsafe(8)
        user.set_password(temp_password)
        user.must_change_password = True
        user.save()
        return Response({
            'temp_password': temp_password,
            'message': f'Contraseña reseteada para {user.rut}. El usuario debe cambiarla al iniciar sesión.'
        })


class ChangePasswordView(views.APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        new_password = request.data.get('new_password')
        if not new_password or len(new_password) < 8:
            return Response({'error': 'La contraseña debe tener al menos 8 caracteres.'}, status=status.HTTP_400_BAD_REQUEST)
        user.set_password(new_password)
        user.must_change_password = False
        user.save()
        return Response({'message': 'Contraseña actualizada correctamente.'})


class RegistroOperativoViewSet(viewsets.ModelViewSet):
    serializer_class = RegistroOperativoSerializer
    permission_classes = [IsAuthenticated, IsPropioUsuarioOJefatura]

    def get_queryset(self):
        user = self.request.user
        from .permissions import ROLES_JEFATURA, _cargo
        if _cargo(user) in ROLES_JEFATURA or user.is_staff:
            return RegistroOperativo.objects.all()
        return RegistroOperativo.objects.filter(rut_trabajador=user)


class AusenciaTemporalViewSet(viewsets.ModelViewSet):
    serializer_class = AusenciaTemporalSerializer
    permission_classes = [IsJefaturaOSuperior]

    def get_queryset(self):
        return AusenciaTemporal.objects.all()


class MeView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        return Response({
            'rut': user.rut,
            'nombre': user.nombre,
            'apellido': user.apellido,
            'cargo': user.cargo,
            'is_active': user.is_active,
            'is_staff': user.is_staff,
            'must_change_password': user.must_change_password,
        })


from rest_framework.parsers import MultiPartParser, FormParser


class CargarTripulacionView(views.APIView):
    """Carga la tripulación desde CSV (RUT;CLAVE;NOMBRE;CARGO). Solo ADMIN."""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        if (getattr(request.user, 'cargo', '') or '').upper() != 'ADMIN':
            return Response({'error': 'Solo el administrador puede cargar la tripulación'}, status=status.HTTP_403_FORBIDDEN)
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'error': 'Adjunte el archivo CSV en el campo "archivo"'}, status=400)
        from apps.operaciones.importadores import importar_tripulacion
        try:
            res = importar_tripulacion(archivo)
        except Exception as e:
            return Response({'error': f'No se pudo procesar el CSV: {e}'}, status=400)
        return Response({'ok': True, **res})


class WipeBaseDatosView(views.APIView):
    """Vacía los datos OPERATIVOS de la base (solo ADMIN). Conserva usuarios y datos maestros.
    Requiere body {confirm: 'ELIMINAR'} para evitar borrados accidentales."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if (getattr(request.user, 'cargo', '') or '').upper() != 'ADMIN':
            return Response({'error': 'Solo el administrador puede vaciar la base de datos'},
                            status=status.HTTP_403_FORBIDDEN)
        if request.data.get('confirm') != 'ELIMINAR':
            return Response({'error': 'Debe confirmar escribiendo ELIMINAR'}, status=400)

        from apps.alertas.models import Emergencia, Incidencia, FallaEquipo
        from apps.bitacora.models import ReporteFinal, NovedadOperativa
        from apps.operaciones.models import (
            ServicioActivo, ServicioHistorico, RegistroEstacion, GraficoMensual,
            ParejaTripulacion, Feriado, ItinerarioEquipo,
        )
        from apps.prevenciones.models import Prevencion

        modelos = [
            ('Emergencias', Emergencia), ('Incidencias', Incidencia), ('Fallas de equipo', FallaEquipo),
            ('Reportes de turno', ReporteFinal), ('Novedades operativas', NovedadOperativa),
            ('Servicios activos', ServicioActivo), ('Servicios históricos', ServicioHistorico),
            ('Registros de estación', RegistroEstacion), ('Gráfico mensual', GraficoMensual),
            ('Parejas de tripulación', ParejaTripulacion), ('Feriados', Feriado),
            ('Itinerario de equipos', ItinerarioEquipo), ('Registro operativo', RegistroOperativo),
            ('Ausencias', AusenciaTemporal), ('Prevenciones', Prevencion),
        ]
        resumen = {}
        total = 0
        with transaction.atomic():
            for nombre, Modelo in modelos:
                n, _ = Modelo.objects.all().delete()
                resumen[nombre] = n
                total += n
        return Response({
            'ok': True,
            'total_eliminados': total,
            'eliminados': resumen,
            'preservados': ['Usuarios', 'Maestro de turnos', 'Itinerario maestro', 'Rutas/estaciones'],
        })


# ─── RECOMENDACIÓN DE REEMPLAZO DE TURNO ─────────────────────────────────────

class RecomendarReemplazoView(views.APIView):
    """Despachador determinista: dado un trabajador ausente y la fecha (+ días),
    busca el turno asignado en el gráfico mensual y recomienda al personal idóneo
    para cubrirlo, optimizando sobretiempo, descanso y horas nocturnas."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from apps.operaciones.models import GraficoMensual, MaestroTurno
        from . import recomendacion as rec

        rut = request.data.get('rut')
        fecha = request.data.get('fecha')
        try:
            dias = max(1, int(request.data.get('dias') or 1))
        except (TypeError, ValueError):
            dias = 1

        if not rut or not fecha:
            return Response({'error': 'rut y fecha son requeridos'}, status=400)

        try:
            ausente = Usuario.objects.get(rut=rut)
        except Usuario.DoesNotExist:
            return Response({'error': 'Trabajador no encontrado'}, status=404)

        cargo_aus = (ausente.cargo or '').upper()
        es_maq = 'MAQUINISTA' in cargo_aus
        es_ay = 'AYUDANTE' in cargo_aus
        if not (es_maq or es_ay):
            return Response({
                'rut': rut,
                'nombre': f"{ausente.nombre} {ausente.apellido}".strip(),
                'cargo': ausente.cargo,
                'recomendaciones': [],
                'mensaje': 'El trabajador no es tripulación (maquinista/ayudante); no requiere reemplazo de turno.',
            })

        try:
            f0 = datetime.date.fromisoformat(fecha)
        except ValueError:
            return Response({'error': 'Formato de fecha inválido (YYYY-MM-DD)'}, status=400)

        recomendaciones = []
        for i in range(min(dias, 14)):
            d = f0 + datetime.timedelta(days=i)
            td = rec.tipo_dia(d)

            graf = GraficoMensual.objects.filter(fecha=d, rut=ausente).first()
            if not graf:
                continue
            num = graf.num_turno.strip()
            if num in ('L', ''):
                continue  # día libre: nada que cubrir

            mt = MaestroTurno.objects.filter(num_turno=num, tipo_dia=td).first()
            if not mt:
                continue
            # La cobertura va de la apertura (abrir turno) al cierre.
            req_ent = (mt.apertura_hora or mt.presentacion_hora or '').strip()
            req_sal = (mt.cierre_hora or '').strip()
            if not rec.hm_valido(req_ent) or not rec.hm_valido(req_sal):
                continue  # sin horario válido para evaluar

            ausentes_d = set(AusenciaTemporal.objects.filter(fecha=d).values_list('rut__rut', flat=True))
            qs = Usuario.objects.filter(is_active=True).exclude(rut=rut)
            if es_maq:
                qs = qs.filter(cargo__icontains='MAQUINISTA')
            else:
                qs = qs.filter(cargo__icontains='AYUDANTE')

            candidatos = []
            for u in qs:
                if u.rut in ausentes_d:
                    continue  # No Aplicable: vacaciones / baja / licencia / ERIC

                og = GraficoMensual.objects.filter(fecha=d, rut=u).first()
                # Sin programación o día libre ('L') = descanso en casa: legalmente NO se puede llamar.
                if not og:
                    continue
                num_u = og.num_turno.strip()
                if num_u in ('L', '') or num_u == num:
                    continue

                omt = MaestroTurno.objects.filter(num_turno=num_u, tipo_dia=td).first()
                if not omt:
                    continue
                serv = (omt.servicios or '').upper()
                if 'DESCANSO' in serv:
                    continue  # turno de descanso: tampoco se llama

                orig_ent = (omt.apertura_hora or omt.presentacion_hora or '').strip()
                orig_sal = (omt.cierre_hora or '').strip()
                if not rec.hm_valido(orig_ent) or not rec.hm_valido(orig_sal):
                    continue

                es_recibidor = 'REC' in serv  # turno de recibidor = disponible presencial
                label = 'su ventana de recibidor' if es_recibidor else 'su turno'
                ev = rec.evaluar(req_ent, req_sal, orig_ent, orig_sal, origen_label=label)
                ev['es_interrupcion'] = not es_recibidor

                candidatos.append({
                    'rut': u.rut,
                    'nombre': f"{u.nombre} {u.apellido}".strip(),
                    'cargo': u.cargo,
                    'tipo': 'RECIBIDOR' if es_recibidor else 'EN_TURNO',
                    'es_recibidor': es_recibidor,
                    'estado': (f'Recibidor turno {num_u} ({orig_ent}–{orig_sal})' if es_recibidor
                               else f'En turno {num_u} ({orig_ent}–{orig_sal})'),
                    'score': round(rec.score(ev), 2),
                    **ev,
                })

            # Prioridad: recibidores (disponibles) primero; los que están en otro turno
            # (interrupción) solo como último recurso. Dentro de cada grupo, menor costo.
            candidatos.sort(key=lambda c: (c['es_interrupcion'], c['score'], c['horas_nocturnas']))

            recomendaciones.append({
                'fecha': str(d),
                'tipo_dia': td,
                'turno': num,
                'servicios': mt.servicios or '',
                'req_entrada': req_ent,
                'req_salida': req_sal,
                'tiene_recibidor': any(not c['es_interrupcion'] for c in candidatos),
                'principal': candidatos[0] if candidatos else None,
                'alternativa': candidatos[1] if len(candidatos) > 1 else None,
                'candidatos': candidatos[:6],
            })

        return Response({
            'rut': rut,
            'nombre': f"{ausente.nombre} {ausente.apellido}".strip(),
            'cargo': ausente.cargo,
            'recomendaciones': recomendaciones,
        })


# ─── HELPERS ────────────────────────────────────────────────────────────────

def _parse_hm(s):
    """Parse 'HH:MM' → datetime.datetime (date=today)"""
    return datetime.datetime.strptime(s.strip(), '%H:%M')

def _horas_entre(t_ini, t_fin):
    """Hours between two datetime objects, advancing t_fin a day if needed."""
    if t_fin < t_ini:
        t_fin += datetime.timedelta(days=1)
    return (t_fin - t_ini).total_seconds() / 3600.0

ROLES_JEFATURA = {'IL', 'INSPECTOR DE LINEA', 'SL', 'SUPERVISOR DE LINEA',
                  'JEFE DE OPERACIONES', 'ADMIN', 'GERENTE', 'GERENCIA'}


# ─── ABRIR TURNO ────────────────────────────────────────────────────────────

class AbrirTurnoView(views.APIView):
    permission_classes = [IsAuthenticated]

    # Minutos de traslado desde cada ubicación hasta EZ/LJ (base)
    TRANSFER_MINUTOS = {
        'CW': 90, 'CC': 45, 'LM': 60, 'OH': 90,
        'HQ': 90, 'GU': 90, 'EZ': 0, 'LJ': 0,
    }

    def post(self, request):
        user = request.user
        cargo = (getattr(user, 'cargo', '') or '').upper().strip()
        if cargo not in ('MAQUINISTA', 'AYUDANTE'):
            return Response({'error': f'Solo tripulación puede abrir turno (cargo recibido: {cargo!r})'}, status=status.HTTP_403_FORBIDDEN)

        fecha_hoy = datetime.date.today().isoformat()

        # Verificar si ya existe un registro para hoy (abierto o cerrado)
        existente = RegistroOperativo.objects.filter(rut_trabajador=user, fecha=fecha_hoy).first()
        if existente:
            if existente.hora_cierre == '':
                return Response({'error': 'Ya tiene un turno abierto hoy'}, status=status.HTTP_400_BAD_REQUEST)
            else:
                # Ya existe un turno cerrado - permitir "reapertura" para modificación
                # Eliminamos el anterior y creamos uno nuevo (o podríamos actualizar)
                return Response({
                    'error': 'Ya cerró turno hoy. Contacte al IL para modificar el registro existente.',
                    'registro_existente': {
                        'id': existente.id,
                        'hora_apertura': existente.hora_apertura,
                        'hora_cierre': existente.hora_cierre,
                        'horas_extras': existente.horas_extras
                    }
                }, status=status.HTTP_400_BAD_REQUEST)

        lugar           = (request.data.get('lugar_apertura') or '').strip().upper()
        hora_presentacion = (request.data.get('hora_presentacion') or '').strip()   # HH:MM ingresada por trabajador
        coincide        = request.data.get('coincide', True)
        motivo          = (request.data.get('motivo') or '').strip()

        # Datos del gráfico mensual (para cálculo de diferencias)
        grafico_apertura_hora = (request.data.get('grafico_apertura_hora') or '').strip()
        grafico_cierre_hora = (request.data.get('grafico_cierre_hora') or '').strip()
        grafico_tiene_descanso_posterior = request.data.get('grafico_tiene_descanso_posterior', False)

        if not lugar or not hora_presentacion:
            return Response({'error': 'lugar_apertura y hora_presentacion son requeridos'}, status=400)

        if lugar not in self.TRANSFER_MINUTOS:
            return Response({'error': f'Lugar inválido. Opciones: {list(self.TRANSFER_MINUTOS)}'}, status=400)

        try:
            t_pres = _parse_hm(hora_presentacion)
        except ValueError:
            return Response({'error': 'Formato de hora inválido. Use HH:MM'}, status=400)

        # Calcular hora de apertura EZ (presentación - traslado)
        transfer = self.TRANSFER_MINUTOS[lugar]
        t_apertura = t_pres - datetime.timedelta(minutes=transfer)
        hora_apertura_ez = t_apertura.strftime('%H:%M')

        # Estado: requiere autorización si no coincide con gráfico
        if not coincide:
            if not motivo:
                return Response({'error': 'Debe ingresar un motivo cuando no coincide con el gráfico'}, status=400)
            estado = 'PENDIENTE'  # Temporal: usar PENDIENTE hasta que migre a PENDIENTE_AUTORIZACION
            obs = f"[PENDIENTE_AUTORIZACION] {motivo}"
        else:
            estado = 'PENDIENTE'
            obs = ''

        reg = RegistroOperativo.objects.create(
            fecha=fecha_hoy,
            rut_trabajador=user,
            lugar_apertura=lugar,
            hora_apertura=hora_apertura_ez,
            inicio_servicio=hora_presentacion,
            hora_cierre='',
            # Datos del gráfico
            grafico_apertura_hora=grafico_apertura_hora,
            grafico_cierre_hora=grafico_cierre_hora,
            grafico_tiene_descanso_posterior=grafico_tiene_descanso_posterior,
            # Inicializar horas extras
            horas_extras=0,
            horas_extras_apertura=0,
            horas_extras_cierre=0,
            horas_extras_7_5=0,
            horas_extras_descanso=0,
            horas_extras_doble_turno=0,
            horas_menos_reposo=0,
            horas_nocturnas=0,
            horas_manejo=0,
            estado=estado,
            observacion_il=obs,
        )
        return Response({
            'id': reg.id,
            'estado': reg.estado,
            'hora_apertura_ez': hora_apertura_ez,
            'hora_presentacion': hora_presentacion,
        })


# ─── CERRAR TURNO ───────────────────────────────────────────────────────────

class CerrarTurnoView(views.APIView):
    """
    Cierre de turno con cálculo completo de horas extras:
    1. Diferencias vs gráfico mensual (apertura anticipada / cierre tardío)
    2. Exceso de 7.5 horas totales
    3. Descanso semanal afectado (post-00:00)
    4. Menos reposo: 10h default, 11.5h si cierre anterior fue en EZ (residencia)
    5. Doble turno / modificación completa
    """
    permission_classes = [IsAuthenticated]

    JORNADA_NORMAL = 7.5  # horas
    REPOSO_MINIMO = 10.0  # horas (default)
    REPOSO_RESIDENCIA_EZ = 11.5  # horas (10h + 1.5h traslado) - solo si cierre fue en EZ

    def post(self, request):
        user = request.user
        cargo = (getattr(user, 'cargo', '') or '').upper().strip()
        if cargo not in ('MAQUINISTA', 'AYUDANTE'):
            return Response({'error': f'Solo tripulación puede cerrar turno (cargo recibido: {cargo!r})'}, status=status.HTTP_403_FORBIDDEN)

        fecha_hoy = datetime.date.today().isoformat()
        try:
            reg = RegistroOperativo.objects.get(rut_trabajador=user, fecha=fecha_hoy, hora_cierre='')
        except RegistroOperativo.DoesNotExist:
            return Response({'error': 'No hay turno abierto para hoy'}, status=404)

        hora_cierre      = (request.data.get('hora_cierre') or '').strip()
        fecha_cierre     = request.data.get('fecha_cierre')  # Opcional: para turnos que cruzan medianoche
        coincide_cierre  = request.data.get('coincide_cierre', True)
        motivo_cierre    = (request.data.get('motivo_cierre') or '').strip()
        inicio_manejo    = (request.data.get('inicio_manejo') or '').strip()
        fin_manejo       = (request.data.get('fin_manejo') or '').strip()

        # Flag para doble turno / modificación completa
        es_doble_turno = request.data.get('es_doble_turno', False)

        if not hora_cierre:
            return Response({'error': 'hora_cierre es requerido'}, status=400)
        if not coincide_cierre and not motivo_cierre:
            return Response({'error': 'Debe ingresar un motivo cuando no coincide con el gráfico'}, status=400)
        if cargo == 'MAQUINISTA' and (not inicio_manejo or not fin_manejo):
            return Response({'error': 'El maquinista debe indicar inicio y fin de manejo'}, status=400)

        fmt = '%H:%M'
        # reg.fecha es un date object, no un string
        fecha_apertura = reg.fecha if isinstance(reg.fecha, datetime.date) else datetime.date.fromisoformat(reg.fecha)
        fecha_cierre_dt = fecha_cierre if fecha_cierre else fecha_hoy

        # Construir datetime completos para manejar turnos que cruzan medianoche
        t_aper = datetime.datetime.combine(
            fecha_apertura,
            datetime.datetime.strptime(reg.hora_apertura, fmt).time()
        )
        t_cierrereal = datetime.datetime.combine(
            datetime.date.fromisoformat(fecha_cierre_dt) if isinstance(fecha_cierre_dt, str) else fecha_cierre_dt,
            datetime.datetime.strptime(hora_cierre, fmt).time()
        )

        # Si no se proporcionó fecha_cierre y el cierre parece anterior a la apertura,
        # asumir que es del día siguiente (turno nocturno)
        if not fecha_cierre and t_cierrereal < t_aper:
            t_cierrereal += datetime.timedelta(days=1)
        elif not fecha_cierre:
            # Validación: si la jornada es menos de 2 horas, probablemente cruzó medianoche
            jornada_temp = (t_cierrereal - t_aper).total_seconds() / 3600
            if jornada_temp < 2.0 and reg.hora_apertura > '12:00' and hora_cierre > '12:00':
                # Ambos son horas PM pero jornada muy corta = cruzó medianoche
                t_cierrereal += datetime.timedelta(days=1)

        # ─────────────────────────────────────────────────────────────────────
        # CÁLCULO 1: HORAS EXTRAS VS GRÁFICO MENSUAL
        # ─────────────────────────────────────────────────────────────────────
        horas_extras_apertura = 0.0  # Por llegar antes de la hora original
        horas_extras_cierre = 0.0    # Por salir después de la hora original
        horas_extras_doble_turno = 0.0  # Por modificación completa

        graf_aper = reg.grafico_apertura_hora
        graf_cier = reg.grafico_cierre_hora

        if graf_aper and graf_cier:
            try:
                t_graf_aper = _parse_hm(graf_aper)
                t_graf_cier = _parse_hm(graf_cier)

                # Normalizar si el turno cruza medianoche
                if t_graf_cier < t_graf_aper:
                    t_graf_cier += datetime.timedelta(days=1)

                if es_doble_turno:
                    # Caso 4: Turno completamente fuera de rango = cobrar totalidad
                    horas_extras_doble_turno = round((t_cierrereal - t_aper).total_seconds() / 3600, 2)
                else:
                    # Casos 1-3: Diferencias parciales
                    # Diferencia en apertura (si abrió antes)
                    if t_aper < t_graf_aper:
                        horas_extras_apertura = round((t_graf_aper - t_aper).total_seconds() / 3600, 2)

                    # Diferencia en cierre (si cerró después)
                    if t_cierrereal > t_graf_cier:
                        horas_extras_cierre = round((t_cierrereal - t_graf_cier).total_seconds() / 3600, 2)
            except ValueError:
                pass  # Si hay error en formato, no calcular diferencias

        # ─────────────────────────────────────────────────────────────────────
        # CÁLCULO 2: HORAS EXTRAS POR EXCEDER 7.5H TOTALES
        # ─────────────────────────────────────────────────────────────────────
        total_jornada = _horas_entre(t_aper, t_cierrereal)
        horas_extras_7_5 = round(max(total_jornada - self.JORNADA_NORMAL, 0.0), 2)

        # ─────────────────────────────────────────────────────────────────────
        # CÁLCULO 3: DESCANSO SEMANAL AFECTADO (post-00:00 cuando debería descansar)
        # ─────────────────────────────────────────────────────────────────────
        horas_extras_descanso = 0.0
        if reg.grafico_tiene_descanso_posterior:
            # Si cierra después de las 00:00 y el día siguiente es descanso
            medianoche = t_aper.replace(hour=0, minute=0) + datetime.timedelta(days=1)
            if t_cierrereal > medianoche:
                minutos_post_00 = (t_cierrereal - medianoche).total_seconds() / 60
                horas_extras_descanso = round(minutos_post_00 / 60, 2)

        # ─────────────────────────────────────────────────────────────────────
        # CÁLCULO 4: MENOS REPOSO (10h default, 11.5h si cierre anterior fue en EZ)
        # ─────────────────────────────────────────────────────────────────────
        horas_menos_reposo = 0.0
        ultimo = RegistroOperativo.objects.filter(
            rut_trabajador=user, hora_cierre__gt=''
        ).exclude(id=reg.id).order_by('-fecha').first()

        if ultimo and ultimo.hora_cierre:
            try:
                # ultimo.fecha y reg.fecha son date objects
                fecha_ult = ultimo.fecha if isinstance(ultimo.fecha, datetime.date) else datetime.date.fromisoformat(ultimo.fecha)
                fecha_reg = reg.fecha if isinstance(reg.fecha, datetime.date) else datetime.date.fromisoformat(reg.fecha)
                dt_ult = datetime.datetime.combine(fecha_ult, datetime.datetime.strptime(ultimo.hora_cierre, fmt).time())
                dt_aper = datetime.datetime.combine(fecha_reg, datetime.datetime.strptime(reg.hora_apertura, fmt).time())
                if dt_aper < dt_ult:
                    dt_aper += datetime.timedelta(days=1)
                h_reposo = (dt_aper - dt_ult).total_seconds() / 3600

                # Si el turno anterior cerró en EZ (residencia laboral), reposo = 11.5h
                # Nota: usamos lugar_apertura del último turno como proxy de lugar_cierre
                ultimo_lugar = (ultimo.lugar_apertura or '').upper()
                if ultimo_lugar == 'EZ':
                    reposo_requerido = self.REPOSO_RESIDENCIA_EZ
                else:
                    reposo_requerido = self.REPOSO_MINIMO

                horas_menos_reposo = round(max(reposo_requerido - h_reposo, 0.0), 2)
            except Exception:
                pass

        # ─────────────────────────────────────────────────────────────────────
        # CÁLCULO 5: HORAS NOCTURNAS (antes de las 07:00)
        # ─────────────────────────────────────────────────────────────────────
        m_noct = 0
        # Usar la misma fecha de apertura para el inicio del servicio
        curr = datetime.datetime.combine(
            fecha_apertura,
            datetime.datetime.strptime(reg.inicio_servicio, fmt).time()
        )
        if curr > t_cierrereal:
            curr -= datetime.timedelta(days=1)
        while curr < t_cierrereal:
            if curr.hour < 7:
                m_noct += 1
            curr += datetime.timedelta(minutes=1)
        horas_nocturnas = round(m_noct / 60.0, 2)

        # ─────────────────────────────────────────────────────────────────────
        # CÁLCULO 6: HORAS DE MANEJO (maquinista)
        # ─────────────────────────────────────────────────────────────────────
        horas_manejo = 0.0
        if cargo == 'MAQUINISTA':
            t_im = _parse_hm(inicio_manejo)
            t_fm = _parse_hm(fin_manejo)
            h_cond = _horas_entre(t_im, t_fm)
            horas_manejo = round(max(h_cond - 5.0, 0.0), 2)

        # ─────────────────────────────────────────────────────────────────────
        # TOTAL DE HORAS EXTRAS (suma de todos los conceptos, sin duplicar)
        # ─────────────────────────────────────────────────────────────────────
        if es_doble_turno:
            # Si es doble turno, se cobra la totalidad del turno, no las diferencias parciales
            horas_extras = horas_extras_doble_turno + horas_extras_descanso + horas_extras_7_5
        else:
            # Suma de diferencias vs gráfico + descanso afectado + exceso 7.5h
            horas_extras = horas_extras_apertura + horas_extras_cierre + horas_extras_descanso + horas_extras_7_5

        horas_extras = round(horas_extras, 2)

        # ─────────────────────────────────────────────────────────────────────
        # ESTADO Y OBSERVACIONES
        # ─────────────────────────────────────────────────────────────────────
        obs_base = reg.observacion_il or ''
        detalle_extras = []
        if horas_extras_apertura > 0:
            detalle_extras.append(f"Apertura anticipada: {horas_extras_apertura}h")
        if horas_extras_cierre > 0:
            detalle_extras.append(f"Cierre tardío: {horas_extras_cierre}h")
        if horas_extras_7_5 > 0:
            detalle_extras.append(f"Exceso 7.5h: {horas_extras_7_5}h")
        if horas_extras_descanso > 0:
            detalle_extras.append(f"Descanso afectado: {horas_extras_descanso}h")
        if horas_extras_doble_turno > 0:
            detalle_extras.append(f"Doble turno/modificación: {horas_extras_doble_turno}h")

        if not coincide_cierre:
            estado_final = 'PENDIENTE'  # Temporal: usar PENDIENTE hasta que migre a PENDIENTE_AUTORIZACION
            obs_final = f"[PENDIENTE_AUTORIZACION] {motivo_cierre}"
            if detalle_extras:
                obs_final += f" | Horas extras: {', '.join(detalle_extras)}"
        elif reg.estado == 'PENDIENTE_AUTORIZACION':
            estado_final = 'PENDIENTE_AUTORIZACION'
            obs_final = obs_base
        else:
            estado_final = 'CONFIRMADO'
            obs_final = obs_base
            if detalle_extras and horas_extras > 0:
                obs_final += f" | Horas extras: {', '.join(detalle_extras)}"

        # ─────────────────────────────────────────────────────────────────────
        # GUARDAR RESULTADOS
        # ─────────────────────────────────────────────────────────────────────
        reg.hora_cierre = hora_cierre
        # Guardar fecha_cierre solo si es diferente a fecha (turno que cruza medianoche)
        fecha_cierre_calc = t_cierrereal.date()
        fecha_reg_cmp = reg.fecha if isinstance(reg.fecha, datetime.date) else datetime.date.fromisoformat(reg.fecha)
        reg.fecha_cierre = fecha_cierre_calc if fecha_cierre_calc != fecha_reg_cmp else None
        reg.horas_extras = horas_extras
        reg.horas_extras_apertura = horas_extras_apertura
        reg.horas_extras_cierre = horas_extras_cierre
        reg.horas_extras_7_5 = horas_extras_7_5
        reg.horas_extras_descanso = horas_extras_descanso
        reg.horas_extras_doble_turno = horas_extras_doble_turno
        reg.horas_nocturnas = horas_nocturnas
        reg.horas_manejo = horas_manejo
        reg.horas_menos_reposo = horas_menos_reposo
        reg.estado = estado_final
        reg.observacion_il = obs_final
        reg.save()

        return Response({
            'id': reg.id,
            'estado': reg.estado,
            'horas_extras': horas_extras,
            'horas_extras_detalle': {
                'vs_grafico_apertura': horas_extras_apertura,
                'vs_grafico_cierre': horas_extras_cierre,
                'exceso_7_5_horas': horas_extras_7_5,
                'descanso_afectado': horas_extras_descanso,
                'doble_turno': horas_extras_doble_turno,
            },
            'horas_nocturnas': horas_nocturnas,
            'horas_manejo': horas_manejo,
            'horas_menos_reposo': horas_menos_reposo,
            'total_jornada': round(total_jornada, 2),
        })


# ─── MI ASISTENCIA HOY ──────────────────────────────────────────────────────

class MiAsistenciaView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        fecha = request.query_params.get('fecha', datetime.date.today().isoformat())
        try:
            reg = RegistroOperativo.objects.get(rut_trabajador=request.user, fecha=fecha)
            return Response(RegistroOperativoSerializer(reg).data)
        except RegistroOperativo.DoesNotExist:
            return Response(None)


# ─── PENDIENTES AUTORIZACIÓN ────────────────────────────────────────────────

from django.db.models import Q

class PendientesAutorizacionView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        cargo = (getattr(request.user, 'cargo', '') or '').upper()
        if cargo not in ROLES_JEFATURA:
            return Response({'error': 'Sin permisos'}, status=403)
        # Buscar registros que requieren autorización (estado explicito o tag en observación)
        pendientes = RegistroOperativo.objects.filter(
            Q(estado='PENDIENTE_AUTORIZACION') |
            Q(observacion_il__contains='[PENDIENTE_AUTORIZACION]')
        ).select_related('rut_trabajador').order_by('-fecha')
        data = []
        for r in pendientes:
            data.append({
                'id': r.id,
                'fecha': str(r.fecha),
                'rut': r.rut_trabajador.rut,
                'nombre': f"{r.rut_trabajador.nombre} {r.rut_trabajador.apellido}".strip(),
                'cargo': r.rut_trabajador.cargo,
                'lugar_apertura': r.lugar_apertura,
                'hora_apertura': r.hora_apertura,
                'inicio_servicio': r.inicio_servicio,
                'hora_cierre': r.hora_cierre,
                'observacion': r.observacion_il,
                'horas_extras': r.horas_extras,
                'horas_nocturnas': r.horas_nocturnas,
                'horas_manejo': r.horas_manejo,
                'horas_menos_reposo': r.horas_menos_reposo,
            })
        return Response(data)

    def post(self, request):
        """Autorizar o rechazar: { id, accion: 'autorizar'|'rechazar' }"""
        cargo = (getattr(request.user, 'cargo', '') or '').upper()
        if cargo not in ROLES_JEFATURA:
            return Response({'error': 'Sin permisos'}, status=403)
        reg_id = request.data.get('id')
        accion = request.data.get('accion', '')

        if not reg_id:
            return Response({'error': 'Se requiere el ID del registro'}, status=400)
        if not accion:
            return Response({'error': 'Se requiere la accion (autorizar o rechazar)'}, status=400)

        try:
            reg = RegistroOperativo.objects.get(id=reg_id)
        except RegistroOperativo.DoesNotExist:
            return Response({'error': 'Registro no encontrado'}, status=404)

        # Verificar que no esté ya procesado
        if reg.estado in ('CONFIRMADO', 'RECHAZADO'):
            return Response({'error': f'El registro ya fue {reg.estado.lower()} anteriormente'}, status=400)

        nombre_il = f"{request.user.nombre} {request.user.apellido}".strip()
        obs_limpia = (reg.observacion_il or '').replace('[PENDIENTE_AUTORIZACION]', '').strip()

        if accion == 'autorizar':
            reg.estado = 'CONFIRMADO'
            reg.observacion_il = f"{obs_limpia} [Autorizado por {nombre_il}]".strip()
        elif accion == 'rechazar':
            reg.estado = 'RECHAZADO'
            reg.observacion_il = f"{obs_limpia} [Rechazado por {nombre_il}]".strip()
        else:
            return Response({'error': 'accion debe ser autorizar o rechazar'}, status=400)

        reg.save()
        return Response({'ok': True, 'estado': reg.estado})


# ─── MI HISTORIAL ASISTENCIA (para tripulación) ─────────────────────────────

class MiHistorialAsistenciaView(views.APIView):
    """Obtener historial personal de asistencia (solo MAQUINISTA/AYUDANTE)"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        cargo = (getattr(request.user, 'cargo', '') or '').upper()
        if cargo not in ('MAQUINISTA', 'AYUDANTE'):
            return Response({'error': 'Solo tripulación puede ver su historial'}, status=403)
        
        fecha_desde = request.query_params.get('desde')
        fecha_hasta = request.query_params.get('hasta')
        
        # Subquery para obtener solo el último registro por fecha
        from django.db.models import Max
        latest_ids = RegistroOperativo.objects.filter(
            rut_trabajador=request.user
        ).values('fecha').annotate(
            max_id=Max('id')
        ).values_list('max_id', flat=True)
        
        qs = RegistroOperativo.objects.filter(
            id__in=latest_ids,
            rut_trabajador=request.user
        ).order_by('fecha')
        
        if fecha_desde:
            qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha__lte=fecha_hasta)
        
        data = []
        for r in qs:
            data.append({
                'id': r.id,
                'fecha': str(r.fecha),
                'lugar_apertura': r.lugar_apertura,
                'hora_apertura': r.hora_apertura,
                'inicio_servicio': r.inicio_servicio,
                'hora_cierre': r.hora_cierre,
                'fecha_cierre': str(r.fecha_cierre) if r.fecha_cierre else None,
                'horas_extras': r.horas_extras,
                'horas_manejo': r.horas_manejo,
                'horas_nocturnas': r.horas_nocturnas,
                'horas_menos_reposo': r.horas_menos_reposo,
                'estado': r.estado,
                'observacion_il': r.observacion_il,
                'modificado_por': None,  # Se llena si fue editado
                'fecha_modificacion': None,
            })
        return Response(data)


# ─── NOTIFICACIONES ──────────────────────────────────────────────────────────

# Notificaciones en memoria (para demo - en producción usar modelo)
_notificaciones_db = {}

def crear_notificacion_modificacion(rut_trabajador, modificador, datos_anteriores, datos_nuevos):
    """Crear notificación cuando un IL modifica horas"""
    notif_id = len(_notificaciones_db.get(rut_trabajador, [])) + 1
    mensaje = f"Tus horas fueron modificadas por {modificador}. "
    mensaje += f"Extras: {datos_anteriores.get('horas_extras', 0)}h → {datos_nuevos.get('horas_extras', 0)}h, "
    mensaje += f"Manejo: {datos_anteriores.get('horas_manejo', 0)}h → {datos_nuevos.get('horas_manejo', 0)}h"
    
    notif = {
        'id': notif_id,
        'tipo': 'modificacion_horas',
        'mensaje': mensaje,
        'fecha': datetime.datetime.now().isoformat(),
        'leida': False,
        'datos': {
            'horas_extras_anterior': datos_anteriores.get('horas_extras'),
            'horas_extras_nueva': datos_nuevos.get('horas_extras'),
            'horas_manejo_anterior': datos_anteriores.get('horas_manejo'),
            'horas_manejo_nueva': datos_nuevos.get('horas_manejo'),
            'modificado_por': modificador,
            'fecha_modificacion': datetime.datetime.now().isoformat(),
        }
    }
    
    if rut_trabajador not in _notificaciones_db:
        _notificaciones_db[rut_trabajador] = []
    _notificaciones_db[rut_trabajador].insert(0, notif)
    return notif


class NotificacionesAsistenciaView(views.APIView):
    """Obtener notificaciones del trabajador"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        cargo = (getattr(request.user, 'cargo', '') or '').upper()
        if cargo not in ('MAQUINISTA', 'AYUDANTE'):
            return Response({'error': 'Sin permisos'}, status=403)
        
        rut = request.user.rut
        notifs = _notificaciones_db.get(rut, [])
        return Response(notifs)

    def post(self, request, pk=None):
        """Marcar notificación como leída"""
        cargo = (getattr(request.user, 'cargo', '') or '').upper()
        if cargo not in ('MAQUINISTA', 'AYUDANTE'):
            return Response({'error': 'Sin permisos'}, status=403)
        
        rut = request.user.rut
        if pk and rut in _notificaciones_db:
            for n in _notificaciones_db[rut]:
                if n['id'] == pk:
                    n['leida'] = True
                    return Response({'ok': True})
        return Response({'error': 'Notificación no encontrada'}, status=404)


# ─── CONSOLIDADO ASISTENCIA ─────────────────────────────────────────────────

class ConsolidadoAsistenciaView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        cargo = (getattr(request.user, 'cargo', '') or '').upper()
        if cargo not in ROLES_JEFATURA | {'JEFE SERVICIO', 'JEFE DE SERVICIO'}:
            return Response({'error': 'Sin permisos'}, status=403)
        fecha_desde = request.query_params.get('desde')
        fecha_hasta = request.query_params.get('hasta')
        rut_filtro = request.query_params.get('rut')
        
        # Subquery para obtener solo el último registro por trabajador por fecha
        from django.db.models import Max
        latest_ids = RegistroOperativo.objects.values('rut_trabajador', 'fecha').annotate(
            max_id=Max('id')
        ).values_list('max_id', flat=True)
        
        qs = RegistroOperativo.objects.filter(id__in=latest_ids).select_related('rut_trabajador').order_by('-fecha')
        
        if fecha_desde:
            qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha__lte=fecha_hasta)
        if rut_filtro:
            qs = qs.filter(rut_trabajador__rut=rut_filtro)
        data = []
        for r in qs:
            data.append({
                'id': r.id,
                'fecha': str(r.fecha),
                'rut': r.rut_trabajador.rut,
                'nombre': f"{r.rut_trabajador.nombre} {r.rut_trabajador.apellido}".strip(),
                'cargo': r.rut_trabajador.cargo,
                'lugar_apertura': r.lugar_apertura,
                'hora_apertura': r.hora_apertura,
                'inicio_servicio': r.inicio_servicio,
                'hora_cierre': r.hora_cierre,
                'fecha_cierre': str(r.fecha_cierre) if r.fecha_cierre else None,
                # Horas extras desglose
                'horas_extras': r.horas_extras,
                'horas_extras_apertura': r.horas_extras_apertura,
                'horas_extras_cierre': r.horas_extras_cierre,
                'horas_extras_7_5': r.horas_extras_7_5,
                'horas_extras_descanso': r.horas_extras_descanso,
                'horas_extras_doble_turno': r.horas_extras_doble_turno,
                'horas_manejo': r.horas_manejo,
                'horas_nocturnas': r.horas_nocturnas,
                'horas_menos_reposo': r.horas_menos_reposo,
                'estado': r.estado,
                'observacion': r.observacion_il,
            })
        return Response(data)


# ─── EDITAR HORAS ───────────────────────────────────────────────────────────

class EditarHorasAsistenciaView(views.APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        cargo = (getattr(request.user, 'cargo', '') or '').upper()
        if cargo not in ROLES_JEFATURA:
            return Response({'error': 'Sin permisos'}, status=403)
        try:
            reg = RegistroOperativo.objects.get(pk=pk)
        except RegistroOperativo.DoesNotExist:
            return Response({'error': 'Registro no encontrado'}, status=404)
        
        # Guardar datos anteriores para notificación
        datos_anteriores = {
            'horas_extras': reg.horas_extras,
            'horas_manejo': reg.horas_manejo,
            'horas_nocturnas': reg.horas_nocturnas,
            'horas_menos_reposo': reg.horas_menos_reposo,
            'hora_apertura': reg.hora_apertura,
            'hora_cierre': reg.hora_cierre,
        }
        
        for field in ('horas_extras', 'horas_manejo', 'horas_nocturnas', 'horas_menos_reposo',
                      'hora_apertura', 'hora_cierre', 'inicio_servicio', 'estado', 'observacion_il'):
            if field in request.data:
                setattr(reg, field, request.data[field])
        reg.save()
        
        # Enviar notificación al trabajador si se modificaron horas
        nombre_modificador = f"{request.user.nombre} {request.user.apellido}".strip()
        crear_notificacion_modificacion(
            rut_trabajador=reg.rut_trabajador.rut,
            modificador=nombre_modificador,
            datos_anteriores=datos_anteriores,
            datos_nuevos=request.data
        )
        
        return Response({
            'ok': True,
            'mensaje': 'Registro actualizado y notificación enviada al trabajador',
            'registro': RegistroOperativoSerializer(reg).data
        })


# ─── ELIMINAR REGISTRO ──────────────────────────────────────────────────────

class EliminarRegistroAsistenciaView(views.APIView):
    """Eliminar registro de asistencia (solo IL/Jefatura)"""
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        cargo = (getattr(request.user, 'cargo', '') or '').upper()
        if cargo not in ROLES_JEFATURA:
            return Response({'error': 'Sin permisos'}, status=403)
        try:
            reg = RegistroOperativo.objects.get(pk=pk)
            reg.delete()
            return Response({'ok': True, 'mensaje': 'Registro eliminado'})
        except RegistroOperativo.DoesNotExist:
            return Response({'error': 'Registro no encontrado'}, status=404)


# ─── EXPORTAR EXCEL ─────────────────────────────────────────────────────────

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse

class ExportarExcelAsistenciaView(views.APIView):
    """Exportar consolidado a Excel con totales por trabajador"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        cargo = (getattr(request.user, 'cargo', '') or '').upper()
        if cargo not in ROLES_JEFATURA | {'JEFE SERVICIO', 'JEFE DE SERVICIO'}:
            return Response({'error': 'Sin permisos'}, status=403)

        fecha_desde = request.query_params.get('desde', '')
        fecha_hasta = request.query_params.get('hasta', '')
        rut_filtro = request.query_params.get('rut', '')
        modo = request.query_params.get('modo', 'detalle')  # 'detalle' o 'resumen'

        # Subquery para obtener solo el último registro por trabajador por fecha (evitar duplicados)
        from django.db.models import Max
        latest_ids = RegistroOperativo.objects.values('rut_trabajador', 'fecha').annotate(
            max_id=Max('id')
        ).values_list('max_id', flat=True)
        
        qs = RegistroOperativo.objects.filter(id__in=latest_ids).select_related('rut_trabajador').order_by('rut_trabajador__rut', 'fecha')
        if fecha_desde:
            qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha__lte=fecha_hasta)
        if rut_filtro:
            qs = qs.filter(rut_trabajador__rut=rut_filtro)

        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencia"

        # Estilos
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        total_fill = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")
        total_font = Font(bold=True)

        if modo == 'resumen':
            # Sheet de resumen por trabajador
            ws.append(['RUT', 'Nombre', 'Cargo', 'Total Días', 'H.Extras Total', 'H.Apertura', 'H.Cierre',
                       'H.Exceso 7.5', 'H.Descanso', 'H.Doble Turno', 'H.Manejo', 'H.Nocturnas', 'H.Menos Reposo'])
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Agrupar por trabajador
            from collections import defaultdict
            resumen = defaultdict(lambda: {
                'nombre': '', 'cargo': '', 'dias': 0,
                'horas_extras': 0, 'horas_extras_apertura': 0, 'horas_extras_cierre': 0,
                'horas_extras_7_5': 0, 'horas_extras_descanso': 0, 'horas_extras_doble_turno': 0,
                'horas_manejo': 0, 'horas_nocturnas': 0, 'horas_menos_reposo': 0
            })

            for r in qs:
                rut = r.rut_trabajador.rut
                resumen[rut]['nombre'] = f"{r.rut_trabajador.nombre} {r.rut_trabajador.apellido}".strip()
                resumen[rut]['cargo'] = r.rut_trabajador.cargo
                resumen[rut]['dias'] += 1
                resumen[rut]['horas_extras'] += r.horas_extras
                resumen[rut]['horas_extras_apertura'] += r.horas_extras_apertura
                resumen[rut]['horas_extras_cierre'] += r.horas_extras_cierre
                resumen[rut]['horas_extras_7_5'] += r.horas_extras_7_5
                resumen[rut]['horas_extras_descanso'] += r.horas_extras_descanso
                resumen[rut]['horas_extras_doble_turno'] += r.horas_extras_doble_turno
                resumen[rut]['horas_manejo'] += r.horas_manejo
                resumen[rut]['horas_nocturnas'] += r.horas_nocturnas
                resumen[rut]['horas_menos_reposo'] += r.horas_menos_reposo

            for rut, datos in sorted(resumen.items()):
                ws.append([rut, datos['nombre'], datos['cargo'], datos['dias'],
                           round(datos['horas_extras'], 2), round(datos['horas_extras_apertura'], 2),
                           round(datos['horas_extras_cierre'], 2), round(datos['horas_extras_7_5'], 2),
                           round(datos['horas_extras_descanso'], 2), round(datos['horas_extras_doble_turno'], 2),
                           round(datos['horas_manejo'], 2), round(datos['horas_nocturnas'], 2),
                           round(datos['horas_menos_reposo'], 2)])

        else:
            # Sheet detallado
            ws.append(['Fecha', 'RUT', 'Nombre', 'Cargo', 'Lugar', 'Apertura', 'Presentación', 'Cierre',
                       'H.Extras Total', 'H.Apertura', 'H.Cierre', 'H.Exceso 7.5', 'H.Descanso', 'H.Doble Turno',
                       'H.Manejo', 'H.Nocturnas', 'H.Menos Reposo', 'Estado', 'Observación'])
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for r in qs:
                ws.append([str(r.fecha), r.rut_trabajador.rut,
                           f"{r.rut_trabajador.nombre} {r.rut_trabajador.apellido}".strip(),
                           r.rut_trabajador.cargo, r.lugar_apertura, r.hora_apertura, r.inicio_servicio,
                           r.hora_cierre or '—', r.horas_extras, r.horas_extras_apertura, r.horas_extras_cierre,
                           r.horas_extras_7_5, r.horas_extras_descanso, r.horas_extras_doble_turno,
                           r.horas_manejo, r.horas_nocturnas, r.horas_menos_reposo, r.estado, r.observacion_il])

        # Ajustar anchos
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Generar nombre de archivo
        filename = f"asistencia_{modo}"
        if fecha_desde and fecha_hasta:
            filename += f"_{fecha_desde}_a_{fecha_hasta}"
        elif fecha_desde:
            filename += f"_desde_{fecha_desde}"
        elif fecha_hasta:
            filename += f"_hasta_{fecha_hasta}"
        if rut_filtro:
            filename += f"_{rut_filtro}"
        filename += ".xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
