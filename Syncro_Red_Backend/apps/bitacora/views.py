from rest_framework import viewsets, decorators
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import ReporteFinal, NovedadOperativa
from .serializers import ReporteFinalSerializer, NovedadOperativaSerializer
from .filters import ReporteFinalFilter
from .services import generar_reporte_turno
from apps.operaciones.models import ServicioActivo, RegistroEstacion


class ReporteFinalViewSet(viewsets.ModelViewSet):
    queryset = ReporteFinal.objects.all().order_by('-creado_en')
    serializer_class = ReporteFinalSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = ReporteFinalFilter

    def perform_create(self, serializer):
        serializer.save(
            usuario=self.request.user,
            cargo=getattr(self.request.user, 'cargo', '')
        )

    @decorators.action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        try:
            fecha = request.query_params.get('fecha')
            print(f"[DEBUG] Exportando Excel para fecha: {fecha}")
            
            reportes = self.queryset
            if fecha:
                reportes = reportes.filter(fecha=fecha)
            
            print(f"[DEBUG] Reportes encontrados: {reportes.count()}")

            wb = Workbook()
            ws = wb.active
            ws.title = 'Reportes'
            ws.append(['Fecha', 'Usuario', 'Cargo', 'Resumen', 'Justificación Cierre', 'Creado'])
            
            for i, rep in enumerate(reportes):
                try:
                    ws.append([
                        str(rep.fecha) if rep.fecha else '',
                        str(rep.usuario) if rep.usuario else '',
                        str(rep.cargo) if rep.cargo else '',
                        rep.resumen_texto[:500] if rep.resumen_texto else '',
                        rep.justificacion_cierre or '',
                        rep.creado_en.isoformat() if rep.creado_en else ''
                    ])
                except Exception as row_error:
                    print(f"[ERROR] Fila {i}: {row_error}")
                    raise

            from io import BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.read(), 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="reportes_{fecha or "todos"}.xlsx"'
            return response
            
        except Exception as e:
            print(f"[ERROR] exportar_excel: {e}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)

    @decorators.action(detail=False, methods=['get'])
    def exportar_bitacora_operativa(self, request):
        """Exporta Excel en formato Bitácora Operativa con detalles de servicios"""
        import logging
        logger = logging.getLogger(__name__)
        
        try:
            fecha = request.query_params.get('fecha')
            logger.info(f"[DEBUG] Exportando Bitácora Operativa para fecha: {fecha}")
            print(f"[DEBUG] Exportando Bitácora Operativa para fecha: {fecha}")
            
            from apps.operaciones.models import ServicioActivo, RegistroEstacion
            logger.info("[DEBUG] Modelos importados correctamente")
            
            # Obtener servicios activos de la fecha
            if fecha:
                servicios = ServicioActivo.objects.filter(fecha=fecha)
            else:
                servicios = ServicioActivo.objects.all()
            
            logger.info(f"[DEBUG] Servicios encontrados: {servicios.count()}")
            print(f"[DEBUG] Servicios encontrados: {servicios.count()}")
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Bitácora Operativa'
            
            # Estilos
            header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            subheader_font = Font(bold=True, size=10)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Título principal
            ws.merge_cells('A1:N1')
            ws['A1'] = 'BITÁCORA OPERATIVA — REGISTRO DE SERVICIOS'
            ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
            ws['A1'].fill = header_fill
            ws['A1'].alignment = center_align
            ws.row_dimensions[1].height = 30
            
            # Info del servicio (primera hoja usa primer servicio como ejemplo)
            primer_servicio = servicios.first()
            if primer_servicio:
                ws['A2'] = 'Fecha:'
                ws['B2'] = str(fecha or primer_servicio.fecha)
                ws['A3'] = 'Equipo:'
                ws['B3'] = primer_servicio.equipo_id or 'Sin equipo'
                ws['A4'] = 'Maquinista:'
                ws['B4'] = primer_servicio.maquinista or 'Sin asignar'
                ws['F2'] = 'Tren:'
                ws['G2'] = primer_servicio.tren_num
                ws['F3'] = 'Estado:'
                ws['G3'] = primer_servicio.estado
                ws['F4'] = 'Ayudante:'
                ws['G4'] = primer_servicio.ayudante or 'Sin asignar'
            
            # Encabezados de columnas (fila 6)
            headers = [
                'Fecha', 'Línea', 'Servicio', 'Equipo', 'Maquinista', 'Ayudante',
                'Estación de Salida', 'Horario', 'Atraso',
                'Estación', 'Horario', 'Atraso', 'Estación de Llegada',
                'Observaciones (motivo del atraso)'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col, value=header)
                cell.font = subheader_font
                cell.fill = subheader_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            # Configurar anchos de columna
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 15
            ws.column_dimensions['K'].width = 10
            ws.column_dimensions['L'].width = 12
            ws.column_dimensions['M'].width = 15
            ws.column_dimensions['N'].width = 40
            
            # Datos por servicio
            row = 7
            RUTAS = {
                1: ['Concepción', 'Talcahuano', 'Hualqui', 'Chiguayante', 'Concepción'],
                2: ['Coronel', 'Concepción', 'Talcahuano', 'Hualqui', 'Chiguayante', 'Concepción', 'Coronel']
            }
            HORARIOS_SERVICIO = {
                20000: ['05:55', '06:10', '06:20', '06:30', '06:40', '06:55', '07:15'],
                20001: ['06:30', '06:45', '06:55', '07:05', '07:15', '07:30', '07:50'],
            }
            
            for servicio in servicios:
                # Usar tren_num para identificar ruta/horarios
                estaciones = RUTAS.get(int(servicio.tren_num) % 1000, ['Origen', 'Destino'])
                horarios = HORARIOS_SERVICIO.get(int(servicio.tren_num), ['', ''])
                
                # Obtener registros de estaciones para este servicio (por tren_num y fecha)
                registros = RegistroEstacion.objects.filter(
                    tren_num=servicio.tren_num,
                    fecha=servicio.fecha
                )
                
                # Primera y última estación con info
                if len(estaciones) >= 2:
                    est_origen = estaciones[0]
                    est_destino = estaciones[-1]
                    horario_salida = horarios[0] if horarios else ''
                    horario_llegada = horarios[-1] if horarios else ''
                else:
                    est_origen = servicio.origen
                    est_destino = servicio.destino
                    horario_salida = ''
                    horario_llegada = ''
                
                # Verificar atrasos (usando estacion_id que contiene el nombre)
                reg_origen = registros.filter(estacion_id__icontains=est_origen).first()
                reg_destino = registros.filter(estacion_id__icontains=est_destino).first()
                
                atraso_salida = 'Sin atraso'
                atraso_llegada = 'Sin atraso'
                obs_salida = ''
                obs_llegada = ''
                
                if reg_origen and reg_origen.estado == 'ATRASO':
                    # Extraer minutos de obs si existe formato "+X min"
                    import re
                    minutos_match = re.search(r'\+(\d+)\s*min', reg_origen.obs or '')
                    minutos = minutos_match.group(1) if minutos_match else ''
                    atraso_salida = f"+{minutos} min" if minutos else 'Con atraso'
                    obs_salida = reg_origen.obs or ''
                    
                if reg_destino and reg_destino.estado == 'ATRASO':
                    import re
                    minutos_match = re.search(r'\+(\d+)\s*min', reg_destino.obs or '')
                    minutos = minutos_match.group(1) if minutos_match else ''
                    atraso_llegada = f"+{minutos} min" if minutos else 'Con atraso'
                    obs_llegada = reg_destino.obs or ''
                
                # Combinar observaciones
                observaciones = []
                if obs_salida:
                    observaciones.append(f"{est_origen}: {obs_salida}")
                if obs_llegada:
                    observaciones.append(f"{est_destino}: {obs_llegada}")
                
                # Agregar atrasos intermedios si existen
                for reg in registros:
                    if reg.estado == 'ATRASO' and reg.estacion_id not in [est_origen, est_destino]:
                        import re
                        minutos_match = re.search(r'\+(\d+)\s*min', reg.obs or '')
                        minutos = minutos_match.group(1) if minutos_match else ''
                        obs = reg.obs or 'Sin detalle'
                        observaciones.append(f"{reg.estacion_id}: Atraso de +{minutos} min — [{obs}]")
                
                ws.cell(row=row, column=1, value=str(servicio.fecha))
                ws.cell(row=row, column=2, value=f"L{int(servicio.tren_num) % 1000}")
                ws.cell(row=row, column=3, value=servicio.tren_num)
                ws.cell(row=row, column=4, value=servicio.equipo_id or 'Sin equipo')
                ws.cell(row=row, column=5, value=servicio.maquinista)
                ws.cell(row=row, column=6, value=servicio.ayudante)
                ws.cell(row=row, column=7, value=est_origen)
                ws.cell(row=row, column=8, value=horario_salida)
                ws.cell(row=row, column=9, value=atraso_salida)
                ws.cell(row=row, column=10, value=est_destino)
                ws.cell(row=row, column=11, value=horario_llegada)
                ws.cell(row=row, column=12, value=atraso_llegada)
                ws.cell(row=row, column=13, value=est_destino)
                ws.cell(row=row, column=14, value=' | '.join(observaciones) if observaciones else '')
                
                # Aplicar bordes
                for col in range(1, 15):
                    ws.cell(row=row, column=col).border = thin_border
                    ws.cell(row=row, column=col).alignment = Alignment(vertical='center', wrap_text=True)
                
                # Colorear celdas de atraso
                if 'Sin atraso' not in atraso_salida:
                    ws.cell(row=row, column=9).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    ws.cell(row=row, column=9).font = Font(color='9C0006')
                else:
                    ws.cell(row=row, column=9).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    ws.cell(row=row, column=9).font = Font(color='006100')
                    
                if 'Sin atraso' not in atraso_llegada:
                    ws.cell(row=row, column=12).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    ws.cell(row=row, column=12).font = Font(color='9C0006')
                else:
                    ws.cell(row=row, column=12).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    ws.cell(row=row, column=12).font = Font(color='006100')
                
                row += 1
            
            from io import BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.read(), 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="Bitacora_Operativa_{fecha or "todos"}.xlsx"'
            return response
            
        except Exception as e:
            print(f"[ERROR] exportar_bitacora_operativa: {e}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)

    @decorators.action(detail=False, methods=['post'])
    def generar_reporte(self, request):
        fecha = request.data.get('fecha')
        if not fecha:
            return Response({'error': 'La fecha es obligatoria'}, status=400)
            
        usuario = request.user
        texto_generado = generar_reporte_turno(fecha, f"{usuario.nombre} {usuario.apellido}", usuario.cargo)
        
        return Response({
            'fecha': fecha,
            'reporte_texto': texto_generado
        })


class NovedadOperativaViewSet(viewsets.ModelViewSet):
    queryset = NovedadOperativa.objects.all().order_by('-timestamp')
    serializer_class = NovedadOperativaSerializer
