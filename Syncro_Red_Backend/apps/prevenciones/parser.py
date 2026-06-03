"""Parser del Boletín de Vía (OIS) — prevenciones de vía desde Excel.

Estructura real del archivo:
  - Encabezado L1: 'NOTIFICACION DE FAENAS EN EL INICIO DEL RECORRIDO (HUALQUI A TALCAHUANO)'.
  - Encabezado L2: 'NOTIFICACION DE FAENAS EN EL INICIO DEL RECORRIDO (CONCEPCIÓN A CORONEL)'.
  - PARADA ESTRICTA: 'PROGRAMACIÓN DE CORTADAS' (todo lo de abajo se ignora).
  - Columnas en dos filas de título (celdas combinadas):
      N° | BLOCK(DESDE,HASTA) | VIA | KM(DESDE,HASTA) | VELOCIDAD Y/O RESTRICCIÓN |
      DESCRIPCIÓN TRABAJOS | HORA(INICIO,TÉRMINO) | FECHA(INICIO,TÉRMINO) | COORDINADO POR
  - La sección L2 NO repite los títulos: usa el mismo mapeo de columnas que L1.

Las columnas se localizan por anclas de palabra clave en la fila de títulos, de modo que
el parser tolera pequeños cambios de orden.
"""

import re
import unicodedata
from openpyxl import load_workbook


def _norm(s):
    s = str(s if s is not None else '')
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()
    return re.sub(r'\s+', ' ', s).strip().upper()


HEADER_L1 = _norm('NOTIFICACION DE FAENAS EN EL INICIO DEL RECORRIDO (HUALQUI A TALCAHUANO)')
HEADER_L2 = _norm('NOTIFICACION DE FAENAS EN EL INICIO DEL RECORRIDO (CONCEPCION A CORONEL)')
STOP = _norm('PROGRAMACION DE CORTADAS')


def _leer_filas(ws):
    return [[('' if c is None else str(c)).strip() for c in row]
            for row in ws.iter_rows(values_only=True)]


def _texto_fila(fila):
    return _norm(' '.join(fila))


def _buscar(filas, aguja, desde=0):
    for i in range(desde, len(filas)):
        if aguja in _texto_fila(filas[i]):
            return i
    return -1


def _hora(v):
    m = re.match(r'^(\d{1,2}):(\d{2})', str(v or '').strip())
    return f'{int(m.group(1)):02d}:{m.group(2)}' if m else ''


def _localizar_columnas(filas, fin_busqueda):
    """Encuentra la fila de títulos y devuelve un dict de índices de columna."""
    for i in range(fin_busqueda):
        t = _texto_fila(filas[i])
        if 'BLOCK' in t and 'KM' in t and ('VELOCID' in t or 'RESTRIC' in t):
            fila = filas[i]
            cols = {}
            for idx, c in enumerate(fila):
                n = _norm(c)
                if not n:
                    continue
                if n in ('BLOCK', 'BLOQUE') and 'block' not in cols:
                    cols['block'] = idx
                elif n == 'VIA' and 'via' not in cols:
                    cols['via'] = idx
                elif n == 'KM' and 'km' not in cols:
                    cols['km'] = idx
                elif ('VELOCID' in n or 'RESTRIC' in n) and 'vel' not in cols:
                    cols['vel'] = idx
                elif ('DESCRIP' in n or 'TRABAJO' in n) and 'desc' not in cols:
                    cols['desc'] = idx
                elif n == 'HORA' and 'hora' not in cols:
                    cols['hora'] = idx
                elif n == 'COORDINADO POR' or n == 'EMPRESA':
                    cols['coord'] = idx
            return i, cols
    return -1, {}


def _g(fila, idx):
    return fila[idx].strip() if idx is not None and 0 <= idx < len(fila) else ''


def _fila_a_prevencion(fila, cols, linea):
    block = cols.get('block')
    km = cols.get('km')
    hora = cols.get('hora')

    desde = _g(fila, block)
    hasta = _g(fila, block + 1) if block is not None else ''
    km_ini = _g(fila, km)
    km_fin = _g(fila, km + 1) if km is not None else ''
    via = _g(fila, cols.get('via'))
    vel = _g(fila, cols.get('vel'))
    desc = _g(fila, cols.get('desc'))
    h_ini = _hora(_g(fila, hora))
    h_ter = _hora(_g(fila, hora + 1)) if hora is not None else ''
    coord = _g(fila, cols.get('coord'))

    if not (desde or km_ini):  # fila sin datos útiles
        return None

    bloque = desde if (desde and desde == hasta) else f'{desde} → {hasta}'.strip(' →')
    return {
        'linea': linea,
        'bloque_afectado': bloque,
        'tramo_desde': desde,
        'tramo_hasta': hasta,
        'via': via,
        'km_inicio': km_ini,
        'km_fin': km_fin,
        'tipo_restriccion': vel,
        'velocidad_restriccion': vel,
        'descripcion_trabajo': desc,
        'causa': desc,
        'hora_inicio': h_ini,
        'hora_termino': h_ter,
        'encargado': coord,
        'texto_original_km': ' | '.join(c for c in fila if c),
        'seccion': 'faena_inicio',
    }


def _extraer(filas, ini, fin, cols, linea):
    out = []
    for i in range(ini, fin):
        texto = _texto_fila(filas[i])
        if STOP in texto:
            break
        if 'NOTIFICACION DE FAENAS' in texto or 'PROGRAMACION' in texto:
            continue  # filas de encabezado de sección
        prev = _fila_a_prevencion(filas[i], cols, linea)
        if prev:
            out.append(prev)
    return out


def parsear_excel(file_obj):
    """Devuelve (lista_prevenciones, advertencias)."""
    wb = load_workbook(file_obj, data_only=True, read_only=True)
    advertencias, todas = [], []

    # Un boletín puede traer varias hojas (ej. 147 y su versión enmendada 147-A).
    # Se usa la ÚLTIMA hoja con secciones de faenas, que es la versión vigente.
    hojas_con_seccion = []
    for ws in wb.worksheets:
        filas = _leer_filas(ws)
        if filas and (_buscar(filas, HEADER_L1) >= 0 or _buscar(filas, HEADER_L2) >= 0):
            hojas_con_seccion.append((ws.title, filas))

    if not hojas_con_seccion:
        wb.close()
        return [], ['No se encontraron los encabezados de faenas (L1/L2) en ninguna hoja.']
    if len(hojas_con_seccion) > 1:
        advertencias.append(
            f"El archivo tiene {len(hojas_con_seccion)} hojas con faenas; se usó la última (vigente): "
            f"'{hojas_con_seccion[-1][0]}'.")

    for titulo, filas in hojas_con_seccion[-1:]:
        stop_idx = _buscar(filas, STOP)
        limite = stop_idx if stop_idx >= 0 else len(filas)

        l1 = _buscar(filas, HEADER_L1)
        l2 = _buscar(filas, HEADER_L2)

        titulos_idx, cols = _localizar_columnas(filas, limite)
        if not cols:
            advertencias.append(f"Hoja '{titulo}': no se encontró la fila de títulos (BLOCK/KM/VELOCIDAD).")
            continue
        # Los datos empiezan tras los títulos (fila de títulos + subtítulo DESDE/HASTA)
        inicio_datos = titulos_idx + 2

        if l1 >= 0:
            fin_l1 = l2 if (l2 > l1) else limite
            todas += _extraer(filas, max(l1 + 1, inicio_datos), fin_l1, cols, 'L1')
        if l2 >= 0:
            todas += _extraer(filas, l2 + 1, limite, cols, 'L2')
    wb.close()
    if not todas and not advertencias:
        advertencias.append('No se extrajeron prevenciones; verifica el formato del archivo.')
    return todas, advertencias
